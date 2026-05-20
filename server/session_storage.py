from __future__ import annotations

import json
from datetime import datetime, date
from io import BytesIO
from pathlib import Path
from typing import Any, Dict, Iterable, List, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

APP_NAME = "Pakar Ternak Nusantara"
STORAGE_VERSION = "1.1"

HEADER_FILL = "166534"
SUBHEADER_FILL = "DCFCE7"
LIGHT_FILL = "F8FAFC"
WARNING_FILL = "FEF3C7"
BORDER_COLOR = "CBD5E1"

PROFILE_LABELS = {
    "farm_name": "Nama Farm / Kelompok",
    "animal_type": "Jenis Ternak",
    "production_goal": "Tujuan Usaha",
    "phase": "Fase Ternak",
    "population": "Populasi (ekor)",
    "average_age": "Umur Rata-rata",
    "average_weight_kg": "Bobot Rata-rata (kg)",
    "location": "Lokasi / Iklim",
    "housing_system": "Sistem Kandang / Kolam",
    "feed_available": "Bahan Pakan Tersedia",
    "water_source": "Sumber Air",
    "main_problem": "Masalah Utama",
    "budget_note": "Catatan Modal / Biaya",
    "market_target": "Target Pasar",
}

PROFILE_KEYS = list(PROFILE_LABELS.keys())
MESSAGE_HEADERS = ["No", "Role", "Isi Pesan"]
RECORD_HEADERS = [
    "Tanggal",
    "Populasi",
    "Bobot Rata-rata (kg)",
    "Pakan Terpakai (kg)",
    "Biaya (Rp)",
    "Mati (ekor)",
    "Telur (butir)",
    "Susu (liter)",
    "Catatan",
    "Perubahan Bobot (kg)",
    "Jarak Hari",
    "ADG (kg/hari)",
    "FCR Estimasi",
]
CALENDAR_HEADERS = ["Tanggal", "Kegiatan", "Deskripsi", "Status", "Catatan"]
HEALTH_HEADERS = ["Kolom", "Isi"]
INSIGHT_HEADERS = ["Kolom", "Isi"]
USAGE_HEADERS = ["Metrik", "Nilai"]
FORMULA_HEADERS = ["No", "Bahan Pakan Terpilih"]


def _json_default(value: Any) -> str:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return str(value)


def _safe_json(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, default=_json_default)


def _parse_json(value: Any, fallback: Any) -> Any:
    if value is None or value == "":
        return fallback
    if isinstance(value, (dict, list)):
        return value
    try:
        return json.loads(str(value))
    except Exception:
        return fallback


def _as_number(value: Any, default: float = 0.0) -> float:
    if value in (None, ""):
        return default
    try:
        return float(value)
    except Exception:
        return default


def _as_int(value: Any, default: int = 0) -> int:
    if value in (None, ""):
        return default
    try:
        return int(float(value))
    except Exception:
        return default


def _parse_date(value: Any) -> Any:
    if isinstance(value, (datetime, date)):
        return value
    if not value:
        return ""
    text = str(value).strip()
    try:
        return datetime.fromisoformat(text).date()
    except Exception:
        return text


def _sheet_title(wb: Workbook, name: str) -> Any:
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False
    return ws


def _style_title(ws, title: str, subtitle: str = "") -> None:
    ws["A1"] = title
    ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=HEADER_FILL)
    ws["A1"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 28
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    if subtitle:
        ws["A2"] = subtitle
        ws["A2"].font = Font(italic=True, color="475569")
        ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)
        ws.row_dimensions[2].height = 34


def _style_table(ws, header_row: int, max_col: int) -> None:
    thin = Side(style="thin", color=BORDER_COLOR)
    for col in range(1, max_col + 1):
        cell = ws.cell(header_row, col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)
    ws.freeze_panes = ws.cell(header_row + 1, 1)


def _style_cells(ws, min_row: int = 1, max_row: int | None = None, max_col: int | None = None) -> None:
    thin = Side(style="thin", color=BORDER_COLOR)
    max_row = max_row or ws.max_row
    max_col = max_col or ws.max_column
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)


def _set_widths(ws, widths: Dict[str, float]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def _auto_filter(ws, header_row: int, max_col: int) -> None:
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(max_col)}{ws.max_row}"


def _write_key_value_sheet(ws, title: str, rows: Iterable[Tuple[str, Any]], subtitle: str = "") -> None:
    _style_title(ws, title, subtitle)
    start = 4
    ws.cell(start, 1, "Kolom")
    ws.cell(start, 2, "Isi")
    _style_table(ws, start, 2)
    for idx, (key, value) in enumerate(rows, start=start + 1):
        ws.cell(idx, 1, key)
        if isinstance(value, (dict, list)):
            ws.cell(idx, 2, _safe_json(value))
        else:
            ws.cell(idx, 2, value)
    _style_cells(ws, start + 1, ws.max_row, 2)
    _set_widths(ws, {"A": 28, "B": 80})


def build_session_payload(
    *,
    session_id: str,
    profile: Dict[str, Any],
    messages: List[Dict[str, Any]],
    records: List[Dict[str, Any]],
    calendar_events: List[Dict[str, Any]],
    last_health_case: Dict[str, Any],
    last_ai_insight: Dict[str, Any],
    formula_selected: List[str],
    usage: Dict[str, Any],
    app_state: Dict[str, Any] | None = None,
) -> Dict[str, Any]:
    now = datetime.now().isoformat(timespec="seconds")
    return {
        "app": APP_NAME,
        "storage_version": STORAGE_VERSION,
        "session_id": session_id,
        "exported_at": now,
        "last_saved_at": now,
        "profile": dict(profile or {}),
        "messages": list(messages or []),
        "records": list(records or []),
        "calendar_events": list(calendar_events or []),
        "last_health_case": dict(last_health_case or {}),
        "last_ai_insight": dict(last_ai_insight or {}),
        "formula_selected": list(formula_selected or []),
        "usage": dict(usage or {}),
        "app_state": dict(app_state or {}),
    }


def session_filename(payload: Dict[str, Any]) -> str:
    profile = payload.get("profile", {}) or {}
    farm_name = str(profile.get("farm_name") or "farm").strip().lower()
    safe_name = "".join(ch if ch.isalnum() else "-" for ch in farm_name).strip("-") or "farm"
    date_code = datetime.now().strftime("%Y%m%d-%H%M")
    session_id = str(payload.get("session_id") or "sesi")[:8]
    return f"pakar-ternak-{safe_name}-{date_code}-{session_id}.xlsx"


def export_session_xlsx(payload: Dict[str, Any], output_path: str | Path | None = None) -> bytes:
    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    profile = payload.get("profile", {}) or {}
    records = sorted(payload.get("records", []) or [], key=lambda r: str(r.get("date", "")))
    messages = payload.get("messages", []) or []
    calendar = sorted(payload.get("calendar_events", []) or [], key=lambda r: str(r.get("date", "")))
    health = payload.get("last_health_case", {}) or {}
    insight = payload.get("last_ai_insight", {}) or {}
    usage = payload.get("usage", {}) or {}
    formula_selected = payload.get("formula_selected", []) or []
    app_state = payload.get("app_state", {}) or {}

    # Ringkasan / dashboard offline
    ws = _sheet_title(wb, "Ringkasan")
    _style_title(
        ws,
        "Pakar Ternak Nusantara - Backup XLSX",
        "File ini dibuat otomatis agar data farm tetap dapat dibaca, disimpan, dan diunggah kembali meskipun sesi Streamlit Online berakhir.",
    )
    summary_rows = [
        ("Aplikasi", payload.get("app", APP_NAME)),
        ("Versi Backup", payload.get("storage_version", STORAGE_VERSION)),
        ("Session ID", payload.get("session_id", "")),
        ("Terakhir Disimpan", payload.get("last_saved_at", payload.get("exported_at", ""))),
        ("Nama Farm", profile.get("farm_name", "")),
        ("Jenis Ternak", profile.get("animal_type", "")),
        ("Tujuan Usaha", profile.get("production_goal", "")),
        ("Fase", profile.get("phase", "")),
        ("Populasi", profile.get("population", "")),
        ("Bobot Rata-rata (kg)", profile.get("average_weight_kg", "")),
        ("Jumlah Catatan Performa", len(records)),
        ("Jumlah Jadwal", len(calendar)),
        ("Jumlah Pesan Chat", len(messages)),
        ("Request AI Sesi", usage.get("requests", 0)),
        ("Estimasi Biaya AI (Rp)", usage.get("estimated_cost_rp", 0)),
        ("Mode Pengguna", app_state.get("user_mode", "")),
        ("Kedalaman Penjelasan", app_state.get("explanation_level", "")),
        ("SOP Terakhir", (app_state.get("last_sop") or {}).get("type", "") if isinstance(app_state.get("last_sop"), dict) else ""),
    ]
    ws.cell(4, 1, "Metrik")
    ws.cell(4, 2, "Nilai")
    _style_table(ws, 4, 2)
    for idx, (k, v) in enumerate(summary_rows, start=5):
        ws.cell(idx, 1, k)
        ws.cell(idx, 2, v)
    ws.cell(25, 1, "Catatan Penting")
    ws.cell(25, 1).font = Font(bold=True)
    ws.cell(26, 1, "1. Simpan file ini di perangkat peternak sebagai backup mandiri.")
    ws.cell(27, 1, "2. Untuk melanjutkan sesi, unggah kembali file ini di menu Backup XLSX pada aplikasi.")
    ws.cell(28, 1, "3. Streamlit Online dapat menghapus file server saat app restart/redeploy; file download ini adalah backup utama.")
    _style_cells(ws, 5, 28, 2)
    _set_widths(ws, {"A": 34, "B": 70})
    ws["B15"].number_format = '#,##0.00'

    # Profil
    ws = _sheet_title(wb, "Profil")
    _style_title(ws, "Profil Peternakan", "Data profil ini dipakai sebagai konteks AI dan dapat diedit/dibaca langsung oleh peternak.")
    ws.cell(4, 1, "Field")
    ws.cell(4, 2, "Nilai")
    ws.cell(4, 3, "Key Sistem")
    _style_table(ws, 4, 3)
    for idx, key in enumerate(PROFILE_KEYS, start=5):
        ws.cell(idx, 1, PROFILE_LABELS[key])
        ws.cell(idx, 2, profile.get(key, ""))
        ws.cell(idx, 3, key)
    _style_cells(ws, 5, ws.max_row, 3)
    _set_widths(ws, {"A": 30, "B": 60, "C": 28})

    # Catatan Performa
    ws = _sheet_title(wb, "Catatan_Performa")
    _style_title(ws, "Catatan Performa", "Berisi data recording harian/mingguan dan kolom analisis sederhana yang bisa dibaca offline.")
    header_row = 4
    for col, header in enumerate(RECORD_HEADERS, start=1):
        ws.cell(header_row, col, header)
    _style_table(ws, header_row, len(RECORD_HEADERS))
    for r_idx, record in enumerate(records, start=header_row + 1):
        ws.cell(r_idx, 1, _parse_date(record.get("date")))
        ws.cell(r_idx, 2, _as_int(record.get("population")))
        ws.cell(r_idx, 3, _as_number(record.get("avg_weight_kg")))
        ws.cell(r_idx, 4, _as_number(record.get("feed_kg")))
        ws.cell(r_idx, 5, _as_number(record.get("cost_rp")))
        ws.cell(r_idx, 6, _as_int(record.get("mortality")))
        ws.cell(r_idx, 7, _as_int(record.get("eggs")))
        ws.cell(r_idx, 8, _as_number(record.get("milk_liter")))
        ws.cell(r_idx, 9, record.get("note", ""))
        if r_idx > header_row + 1:
            ws.cell(r_idx, 10, f"=IFERROR(C{r_idx}-C{r_idx-1},\"\")")
            ws.cell(r_idx, 11, f"=IFERROR(A{r_idx}-A{r_idx-1},\"\")")
            ws.cell(r_idx, 12, f"=IFERROR(J{r_idx}/K{r_idx},\"\")")
            ws.cell(r_idx, 13, f"=IFERROR(D{r_idx}/(J{r_idx}*B{r_idx}),\"\")")
    if not records:
        ws.cell(header_row + 1, 1, "Belum ada catatan performa.")
    _style_cells(ws, header_row + 1, max(ws.max_row, header_row + 1), len(RECORD_HEADERS))
    _auto_filter(ws, header_row, len(RECORD_HEADERS))
    _set_widths(ws, {"A": 14, "B": 12, "C": 18, "D": 18, "E": 16, "F": 12, "G": 12, "H": 13, "I": 40, "J": 18, "K": 12, "L": 16, "M": 16})
    for row in range(header_row + 1, ws.max_row + 1):
        ws.cell(row, 1).number_format = "yyyy-mm-dd"
        ws.cell(row, 3).number_format = "0.00"
        ws.cell(row, 4).number_format = "0.00"
        ws.cell(row, 5).number_format = '#,##0'
        ws.cell(row, 8).number_format = "0.00"
        ws.cell(row, 10).number_format = "0.00"
        ws.cell(row, 12).number_format = "0.000"
        ws.cell(row, 13).number_format = "0.00"

    # Kalender
    ws = _sheet_title(wb, "Kalender")
    _style_title(ws, "Kalender Manajemen", "Jadwal sanitasi, recording, evaluasi pakan, reproduksi, kesehatan, dan kegiatan lain.")
    header_row = 4
    for col, header in enumerate(CALENDAR_HEADERS, start=1):
        ws.cell(header_row, col, header)
    _style_table(ws, header_row, len(CALENDAR_HEADERS))
    for r_idx, item in enumerate(calendar, start=header_row + 1):
        ws.cell(r_idx, 1, _parse_date(item.get("date")))
        ws.cell(r_idx, 2, item.get("title", ""))
        ws.cell(r_idx, 3, item.get("description", ""))
        ws.cell(r_idx, 4, item.get("status", "Belum dicek"))
        ws.cell(r_idx, 5, item.get("note", ""))
    if not calendar:
        ws.cell(header_row + 1, 1, "Belum ada jadwal.")
    _style_cells(ws, header_row + 1, max(ws.max_row, header_row + 1), len(CALENDAR_HEADERS))
    _auto_filter(ws, header_row, len(CALENDAR_HEADERS))
    _set_widths(ws, {"A": 14, "B": 32, "C": 70, "D": 18, "E": 35})
    for row in range(header_row + 1, ws.max_row + 1):
        ws.cell(row, 1).number_format = "yyyy-mm-dd"

    # Chat
    ws = _sheet_title(wb, "Chat")
    _style_title(ws, "Riwayat Chat", "Percakapan pengguna dan asisten. Data ini bisa dipulihkan kembali ke aplikasi.")
    header_row = 4
    for col, header in enumerate(MESSAGE_HEADERS, start=1):
        ws.cell(header_row, col, header)
    _style_table(ws, header_row, len(MESSAGE_HEADERS))
    for r_idx, msg in enumerate(messages, start=header_row + 1):
        ws.cell(r_idx, 1, r_idx - header_row)
        ws.cell(r_idx, 2, msg.get("role", ""))
        ws.cell(r_idx, 3, msg.get("content", ""))
    if not messages:
        ws.cell(header_row + 1, 1, "Belum ada riwayat chat.")
    _style_cells(ws, header_row + 1, max(ws.max_row, header_row + 1), len(MESSAGE_HEADERS))
    _set_widths(ws, {"A": 8, "B": 14, "C": 100})

    # Kesehatan
    ws = _sheet_title(wb, "Kesehatan")
    _write_key_value_sheet(ws, "Kasus Kesehatan Terakhir", [(k, v) for k, v in health.items()], "Data triase terakhir yang pernah dianalisis.")

    # Insight AI
    ws = _sheet_title(wb, "Insight_AI")
    insight_rows = [("generated_at", insight.get("generated_at", "")), ("content", insight.get("content", "")), ("scorecard", insight.get("scorecard", {})), ("meta", insight.get("meta", {}))]
    _write_key_value_sheet(ws, "Insight AI Terakhir", insight_rows, "Insight terakhir yang dibuat AI berdasarkan data farm.")

    # Pakan
    ws = _sheet_title(wb, "Pakan")
    _style_title(ws, "Bahan Pakan Terpilih", "Daftar bahan yang terakhir dipilih dalam modul formulasi pakan.")
    header_row = 4
    for col, header in enumerate(FORMULA_HEADERS, start=1):
        ws.cell(header_row, col, header)
    _style_table(ws, header_row, len(FORMULA_HEADERS))
    for r_idx, name in enumerate(formula_selected, start=header_row + 1):
        ws.cell(r_idx, 1, r_idx - header_row)
        ws.cell(r_idx, 2, name)
    if not formula_selected:
        ws.cell(header_row + 1, 1, "Belum ada bahan pakan terpilih.")
    _style_cells(ws, header_row + 1, max(ws.max_row, header_row + 1), len(FORMULA_HEADERS))
    _set_widths(ws, {"A": 8, "B": 42})

    # Pengaturan dan fitur lanjutan
    ws = _sheet_title(wb, "Pengaturan")
    state_rows = []
    for key, value in app_state.items():
        state_rows.append((key, value))
    _write_key_value_sheet(ws, "Pengaturan dan Fitur Lanjutan", state_rows, "Mode pengguna, konsultasi bertahap, SOP, prediksi, biosecurity, dan progress edukasi.")

    # SOP terakhir
    last_sop = app_state.get("last_sop", {}) if isinstance(app_state, dict) else {}
    ws = _sheet_title(wb, "SOP_Terakhir")
    sop_rows = [(k, v) for k, v in (last_sop or {}).items()] if isinstance(last_sop, dict) else [("content", str(last_sop))]
    _write_key_value_sheet(ws, "SOP Terakhir", sop_rows, "SOP terakhir yang dibuat/disimpan dalam aplikasi.")

    # Prediksi usaha terakhir
    prediction = app_state.get("last_prediction", {}) if isinstance(app_state, dict) else {}
    ws = _sheet_title(wb, "Prediksi_Usaha")
    pred_rows = [(k, v) for k, v in (prediction or {}).items()] if isinstance(prediction, dict) else [("content", str(prediction))]
    _write_key_value_sheet(ws, "Prediksi Usaha Terakhir", pred_rows, "Prediksi stok pakan, panen, dan estimasi usaha terakhir.")

    # Pemakaian AI
    ws = _sheet_title(wb, "Pemakaian_AI")
    _write_key_value_sheet(ws, "Pemakaian AI Sesi", [(k, v) for k, v in usage.items()], "Estimasi pemakaian AI selama sesi aktif.")

    # Raw JSON for robust restore
    ws = _sheet_title(wb, "RAW_JSON")
    _style_title(ws, "RAW JSON", "Sheet ini dipakai sistem untuk pemulihan data yang lebih akurat. Jangan dihapus jika file akan diunggah kembali.")
    ws["A4"] = "payload_json"
    ws["A4"].font = Font(bold=True, color="FFFFFF")
    ws["A4"].fill = PatternFill("solid", fgColor=HEADER_FILL)
    ws["A5"] = _safe_json(payload)
    ws["A5"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 120
    ws.row_dimensions[5].height = 120
    ws.sheet_state = "hidden"

    # Workbook-level polish
    for ws in wb.worksheets:
        if ws.title != "RAW_JSON":
            for row in ws.iter_rows():
                for cell in row:
                    if cell.row > 1 and cell.value not in (None, ""):
                        cell.alignment = Alignment(vertical="top", wrap_text=True)
            for row_idx in range(1, min(ws.max_row, 120) + 1):
                ws.row_dimensions[row_idx].height = min(max(ws.row_dimensions[row_idx].height or 18, 18), 70)

    bio = BytesIO()
    wb.save(bio)
    data = bio.getvalue()
    if output_path:
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        Path(output_path).write_bytes(data)
    return data


def import_session_xlsx(file_or_bytes: Any) -> Dict[str, Any]:
    if hasattr(file_or_bytes, "read"):
        data = file_or_bytes.read()
    elif isinstance(file_or_bytes, (bytes, bytearray)):
        data = bytes(file_or_bytes)
    else:
        data = Path(file_or_bytes).read_bytes()
    wb = load_workbook(BytesIO(data), data_only=True)

    if "RAW_JSON" in wb.sheetnames:
        raw = wb["RAW_JSON"]["A5"].value
        payload = _parse_json(raw, {})
        if isinstance(payload, dict) and payload.get("app"):
            return _normalise_imported_payload(payload)

    # Fallback if RAW_JSON is missing; reconstruct from human-readable sheets.
    payload: Dict[str, Any] = {
        "app": APP_NAME,
        "storage_version": STORAGE_VERSION,
        "session_id": "imported",
        "exported_at": datetime.now().isoformat(timespec="seconds"),
        "profile": {},
        "messages": [],
        "records": [],
        "calendar_events": [],
        "last_health_case": {},
        "last_ai_insight": {},
        "formula_selected": [],
        "usage": {},
        "app_state": {},
    }

    if "Profil" in wb.sheetnames:
        ws = wb["Profil"]
        for row in range(5, ws.max_row + 1):
            key = ws.cell(row, 3).value
            value = ws.cell(row, 2).value
            if key:
                payload["profile"][str(key)] = value if value is not None else ""

    if "Chat" in wb.sheetnames:
        ws = wb["Chat"]
        for row in range(5, ws.max_row + 1):
            role = ws.cell(row, 2).value
            content = ws.cell(row, 3).value
            if role in {"user", "assistant", "system"} and content:
                payload["messages"].append({"role": str(role), "content": str(content)})

    if "Catatan_Performa" in wb.sheetnames:
        ws = wb["Catatan_Performa"]
        for row in range(5, ws.max_row + 1):
            first = ws.cell(row, 1).value
            if not first or str(first).startswith("Belum"):
                continue
            date_val = first.isoformat() if isinstance(first, (datetime, date)) else str(first)
            payload["records"].append({
                "date": date_val,
                "population": _as_int(ws.cell(row, 2).value),
                "avg_weight_kg": _as_number(ws.cell(row, 3).value),
                "feed_kg": _as_number(ws.cell(row, 4).value),
                "cost_rp": _as_number(ws.cell(row, 5).value),
                "mortality": _as_int(ws.cell(row, 6).value),
                "eggs": _as_int(ws.cell(row, 7).value),
                "milk_liter": _as_number(ws.cell(row, 8).value),
                "note": str(ws.cell(row, 9).value or ""),
            })

    if "Kalender" in wb.sheetnames:
        ws = wb["Kalender"]
        for row in range(5, ws.max_row + 1):
            first = ws.cell(row, 1).value
            if not first or str(first).startswith("Belum"):
                continue
            date_val = first.isoformat() if isinstance(first, (datetime, date)) else str(first)
            payload["calendar_events"].append({
                "date": date_val,
                "title": str(ws.cell(row, 2).value or ""),
                "description": str(ws.cell(row, 3).value or ""),
                "status": str(ws.cell(row, 4).value or ""),
                "note": str(ws.cell(row, 5).value or ""),
            })

    if "Kesehatan" in wb.sheetnames:
        ws = wb["Kesehatan"]
        for row in range(5, ws.max_row + 1):
            key = ws.cell(row, 1).value
            value = ws.cell(row, 2).value
            if key:
                payload["last_health_case"][str(key)] = value if value is not None else ""

    if "Insight_AI" in wb.sheetnames:
        ws = wb["Insight_AI"]
        for row in range(5, ws.max_row + 1):
            key = ws.cell(row, 1).value
            value = ws.cell(row, 2).value
            if key == "scorecard" or key == "meta":
                payload["last_ai_insight"][str(key)] = _parse_json(value, {})
            elif key:
                payload["last_ai_insight"][str(key)] = value if value is not None else ""

    if "Pakan" in wb.sheetnames:
        ws = wb["Pakan"]
        for row in range(5, ws.max_row + 1):
            value = ws.cell(row, 2).value
            if value and not str(value).startswith("Belum"):
                payload["formula_selected"].append(str(value))

    if "Pemakaian_AI" in wb.sheetnames:
        ws = wb["Pemakaian_AI"]
        for row in range(5, ws.max_row + 1):
            key = ws.cell(row, 1).value
            value = ws.cell(row, 2).value
            if key:
                payload["usage"][str(key)] = value

    return _normalise_imported_payload(payload)


def _normalise_imported_payload(payload: Dict[str, Any]) -> Dict[str, Any]:
    payload = dict(payload or {})
    payload.setdefault("app", APP_NAME)
    payload.setdefault("storage_version", STORAGE_VERSION)
    payload.setdefault("session_id", "imported")
    payload.setdefault("profile", {})
    payload.setdefault("messages", [])
    payload.setdefault("records", [])
    payload.setdefault("calendar_events", [])
    payload.setdefault("last_health_case", {})
    payload.setdefault("last_ai_insight", {})
    payload.setdefault("formula_selected", [])
    payload.setdefault("usage", {})
    payload.setdefault("app_state", {})
    if not isinstance(payload["messages"], list):
        payload["messages"] = []
    if not isinstance(payload["records"], list):
        payload["records"] = []
    if not isinstance(payload["calendar_events"], list):
        payload["calendar_events"] = []
    if not isinstance(payload["formula_selected"], list):
        payload["formula_selected"] = []
    if not isinstance(payload["profile"], dict):
        payload["profile"] = {}
    if not isinstance(payload["last_health_case"], dict):
        payload["last_health_case"] = {}
    if not isinstance(payload["last_ai_insight"], dict):
        payload["last_ai_insight"] = {}
    if not isinstance(payload["usage"], dict):
        payload["usage"] = {}
    if not isinstance(payload["app_state"], dict):
        payload["app_state"] = {}
    return payload
